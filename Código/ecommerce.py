#!/usr/bin/env python
# coding: utf-8

# In[6]:


from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import tkinter as tk
from datetime import datetime
import sys, os
import pandas as pd
import openpyxl as ox
import datetime
import threading
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfilename
import traceback
import configparser
import openpyxl
import traceback
import numpy as np
from itertools import product
from sys import stdout as out
from mip import Model, xsum, minimize, BINARY
import operator


# In[7]:


class MenuApp(tk.Frame):

    def __init__(self, master=None):
        
        dirpath = os.getcwd()
        if(not os.path.isfile(dirpath + '\\Images\\fundo1.png')):
            self.msgInfo( "O arquivo fundo1.png não encontrado na pasta Images")
            return
        
        tk.Frame.__init__(self, master)
        self.grid()
        self.master.withdraw()
        self.master.resizable(width=False, height=False)
        self.master.protocol('WM_DELETE_WINDOW', self.exit_app)


        self.local_path = os.getcwd()

        self.images_path = f"{self.local_path}\\Images\\"

        self.arqLayout = ""



    def exit_app(self):

        self.master.destroy()
        return
        #self.ExcelApp.Quit()

    def msgInfo(self, messagem):

        msg_window = Toplevel()
        msg_window.title("Ecommerce - Gestão de rotas")
        msg_window.attributes('-topmost', True)
        msgbox = tk.messagebox.showwarning('Ecommerce - Gestão de rotas', messagem, parent=msg_window)
        msg_window.destroy()

        return


    def main_screen(self):

        def on_resize(event):
            image = bg_image.resize((event.width, event.height), Image.ANTIALIAS)
            label_img.image = ImageTk.PhotoImage(image)
            label_img.config(image=label_img.image)

        
        top = tk.Toplevel(self.master)
        top.title(f"Ecommerce - Gestão de rotas")

        top.geometry("500x300")
        top.configure(background="blue")
        top.protocol('WM_DELETE_WINDOW', self.exit_app)

        # BACKGROUND COLOR
        bg_color = "#1f5e9c"#"#396691"

        # CONFIG TOP E FRAMES
        for r in range(30):
            top.rowconfigure(r, weight=1)
        for c in range(7):
            top.columnconfigure(c, weight=1)

        mainframe = tk.Frame(top, bg="blue")
        mainframe.grid(row=0, column=0, rowspan=28, columnspan=7, sticky=tk.W+tk.E+tk.N+tk.S)
        configframe = tk.Frame(top, bg=bg_color)
        configframe.grid(row=28, column=0, rowspan=2, columnspan=7, padx=5, sticky=tk.W+tk.E+tk.N+tk.S)

        for r in range(9):
            configframe.rowconfigure(r, weight=1)
        for c in range(7):
            configframe.columnconfigure(c, weight=1)

        # IMAGEM BACKGROUND
        bg_image = Image.open(f'{self.images_path}\\fundo1.png')
        label_img = tk.Label(mainframe, image=ImageTk.PhotoImage(bg_image), highlightthickness=0, borderwidth=0)
        label_img.place(x=-10, y=-20, relwidth=1.03, relheight=1.06)
        label_img.bind('<Configure>', on_resize)

        # BOTAO 1
        btn_image1 = ImageTk.PhotoImage(Image.open(f'{self.images_path}botao1.png'))
        btn1 = tk.Button(top, image=btn_image1, command=lambda: self.recon_button(1))
        btn1.config(width=120, height=65)
        btn1.grid(row=14, column=3, pady=2)


        # USERNAME E PASSWORD E CONN
        lbl_user = tk.Label(configframe, bg=bg_color, text="Favor configurar as informações no excel antes de rodar")
        lbl_user.config(font=("Arial", 12))
        lbl_user.grid(row=1, column=0, sticky=tk.E)

        
        self.master.mainloop()

    
    def select_arq(self, titulo, originPath):
        
        msg_window = Toplevel()
        msg_window.title("Ecommerce - Gestão de rotas")
        msg_window.attributes('-topmost', True)
        file = tk.filedialog.askopenfilename(initialdir=originPath,
                                          filetypes=(("Excel File", "*.xlsx"), ("Excel File", "*.xlsm"), ("All Files", "*.*")),
                                          title=titulo, parent=msg_window)
        msg_window.destroy()
        
        return file
        


    def recon_button(self, funct):   
        
        
        if self.arqLayout == "":
            originPath = os.getcwd()
            if os.path.isfile(originPath + '\\Ecommerce - Layout.xlsx'):
                self.arqLayout = originPath + '\\Ecommerce - Layout.xlsx'
            else:
                self.arqLayout = self.select_arq("Selecione o arquvio Ecommerce - Layout", originPath)

            
    
        if self.arqLayout == "":
            return
        
        if funct == 1:
            self.otimizacao(self.arqLayout)

    
    def otimizacao(self, nomeanexo):
    
        #Carregar dados
        workbookLayout = openpyxl.load_workbook(nomeanexo , data_only=True)
        worksheetConfig = workbookLayout['Config']
        worksheetDados = workbookLayout['Distancia']

        limiteTempo = worksheetConfig.cell(row=3, column=7).value
        if limiteTempo == None:
            limiteTempo = 60

        linha = 4
        cont = worksheetConfig.cell(row=3, column=2).value
        disctPrio = {}        

        #Pegar os dados das Prioridades das cidades
        while( worksheetConfig.cell(row=linha, column=4).value != None):
            cidade = worksheetConfig.cell(row=linha, column=4).value
            prior = worksheetConfig.cell(row=linha, column=5).value
            if(prior == 0 or prior == None or cidade == None ):
                self.msgInfo("Existem prioridades/cidades vazias ou com valor 0")
                return
            disctPrio[cidade] = prior
            cont = cont - 1
            linha = linha + 1
        
        
        dictPos = {}
        listaCidades = []
        listaPrioridades = []
        listaPrio = []
        
        #ordenar a lista de Prioridades de forma crescente e armazenar nos respectivos vetores 
        #dictPos -> dicionario para saber qual a posicao da cidade no vetor listaCidades/listaPrioridades
        #listaCidades -> vetor que armazena os nomes das cidades
        #listaPrioridades -> vetor que armazena a prioridade de cada cidade
        listaPrio = sorted(disctPrio.items(), key=operator.itemgetter(1))
        contPos = 0
        for cidade,prior in listaPrio:
            dictPos[cidade] = contPos
            listaCidades.append(cidade)
            listaPrioridades.append(prior)
            contPos = contPos + 1
        
        #Colocar distancia "INFINITA" na matriz de custo inicial 
        #c -> matriz de distancia com prioridade
        #distOrig -> matriz de distancia sem prioridade
        c = []
        for origem in listaCidades:
            listaA = []
            for destino in listaCidades:
                listaA.append(999999)
            c.append(listaA)

        distOrig = []
        for origem in listaCidades:
            listaA = []
            for destino in listaCidades:
                listaA.append(999999)
            distOrig.append(listaA)

        if cont != 0:
            self.msgInfo("Verifique se a lista de prioridades condiz com a quantidade de locias escolhida")
            return 

        #Setar a matriz de distancia de acordo com o que foi preenchido na planilha
        linha = 2
        while( worksheetDados.cell(row=linha, column=2).value != None):
            origem = worksheetDados.cell(row=linha, column=1).value
            destino = worksheetDados.cell(row=linha, column=2).value
            distancia = worksheetDados.cell(row=linha, column=3).value
            posO = dictPos.get(origem)
            posD = dictPos.get(destino)
            if posO == None or posD == None :
                self.msgInfo("Existem locais não cadastrados em prioridades nas Origens e/ou Destinos")
                return

            prioridade = 1 / (listaPrioridades[posD])
            c[posO][posD] = (distancia * prioridade)
            distOrig[posO][posD] = distancia 
            if c[posD][posO] == 999999:
                prioridade = 1 / (listaPrioridades[posD]) 
                c[posD][posO] = (distancia * prioridade)
                distOrig[posD][posO] = distancia 
            linha = linha + 1
        
        # número de nós e lista de vértices
        n, V = len(c), set(range(len(c)))

        model = Model()

        # variáveis binárias indicando se arco (i,j) é usado na rota
        x = [[model.add_var(var_type=BINARY) for j in V] for i in V]

        # variáveis contínuas para prevenção de sub-rotas: cada cidade terá
        # um identificador numérico maior na rota, excetuando-se a primeira
        y = [model.add_var() for i in V]

        # função objetivo: minimizar custo total
        model.objective = minimize(xsum(c[i][j]*x[i][j] for i in V for j in V))

        # restrição : sair de cada cidade somente uma vez
        for i in V:
            model += xsum(x[i][j] for j in V - {i}) == 1

        # restrição : entrar em cada cidade somente uma vez
        for i in V:
            model += xsum(x[j][i] for j in V - {i}) == 1

        # eliminação de sub-rotas
        for (i, j) in product(V - {0}, V - {0}):
            if i != j:
                model += y[i] - (n+1)*x[i][j] >= y[j]-n

        # otimização com limite de tempo de 30 segundos
        model.optimize(max_seconds=limiteTempo)
        listaResult = []
        resul = ""
        # verificando se ao menos uma solução válida foi encontrada
        if model.num_solutions:
            #print(model.objective_value)
            #out.write('route with total distance %g found: %s'
            #    % (model.objective_value, listaCidades[0]))
            listaResult.append( listaCidades[0])
            resul += listaCidades[0]
            nc = 0
            while True:
                nc = [i for i in V if x[nc][i].x >= 0.99][0]
                #out.write(' -> %s' % listaCidades[nc])
                resul += "-> " + listaCidades[nc]
                listaResult.append( listaCidades[nc])
                if nc == 0:
                    break
            #out.write('\n')

        distResul = 0
        for cidade in range(1,len(listaResult)):
            origem = listaResult[cidade-1]
            destino = listaResult[cidade]
            #print(origem,destino)
            posO = dictPos.get(origem)
            posD = dictPos.get(destino)
            if distOrig[posO][posD] == 999999:
                worksheetConfig.cell(row=12, column=10).value = "Caminho não encontrado"
                worksheetConfig.cell(row=12, column=11).value = ""
                workbookLayout.save(nomeanexo)
                self.msgInfo("Calculo de rota finalizado. Confira a respota na aba de Configuração")
                return

            distResul += distOrig[posO][posD]

        worksheetConfig.cell(row=12, column=10).value = resul
        worksheetConfig.cell(row=12, column=11).value = distResul
        workbookLayout.save(nomeanexo)
        self.msgInfo("Calculo de rota finalizado. Confira a respota na aba de Configuração")
        return
    
class FileDialog:
    @classmethod
    def open_file(cls, dir_path):
        filename = askopenfilename(initialdir=dir_path,
                                   filetypes=(("Excel File", "*.xlsx"), ("All Files", "*.*")),
                                   title="Choose a file.")
        #print(filename)

        if not os.path.isfile(filename):
            return None

        return filename


class Th(threading.Thread):

    def __init__(self, script, user, passw, conn, companies, funct, file=None, desired_date=None, posting_date=None):
        threading.Thread.__init__(self, daemon=True)
        self.script = script
        self.user = user
        self.passw = passw
        self.conn = conn
        self.companies = companies
        self.funct = funct
        self.file = file
        self.desired_date = desired_date
        self.posting_date = posting_date
        self.stop = threading.Event()

    def run(self):
        pass



if __name__ == "__main__":

    try:
        app = MenuApp()
        app.main_screen()

    except:
        print(traceback.format_exc())
        input("PRESSIONE ENTER PARA ENCERRAR")




# In[ ]:




